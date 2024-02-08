import sys
import pandas as pd
import tkinter as tk
from tkinter.messagebox import showinfo
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def info_window(title, text):
    showinfo(title, text)

def get_file_path():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title="Excel-Datei auswählen")
    return file_path

def convert_to_xlsx(file_path):
    df = pd.read_excel(file_path)  # Read the .xls file into a DataFrame
    df.to_excel(file_path + 'x', index=False)  # Save the DataFrame to a new .xlsx file
    return file_path + 'x'

def color_overlapping_numbers(excel_file_path, common_numbers):
    wb = load_workbook(excel_file_path)    # Load the Excel file using Engine
    ws = wb.active  # Access the active sheet
    col_letter = ws.cell(row=1, column=1).column_letter   # Get the column letter for the first column
    for idx, value in enumerate(pd.read_excel(excel_file_path)[pd.read_excel(excel_file_path).columns[0]].isin(common_numbers)):  # Set the fill color for overlapping numbers
        if value:
            cell = ws[col_letter + str(idx + 2)]  # Adding 2 because Excel sheet indexing starts from 1
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow fill color
    wb.save(excel_file_path) # Save the modified workbook

# first upload prompt
showinfo("Warnung","Alle angegebenen Dateien mit der Endung xlsx werden direkt makiert!!!\nAlle xls Dateien werden als neue xlsx Datei generiert, die xls bleibt erhalten!")
excel_file1_path = get_file_path()
if excel_file1_path.endswith(".xls") or excel_file1_path.endswith(".xlsx"):
    info_window("Erfolg", "Die erste Excel-Datei wurde angenommen!")
else:
    info_window("Fehlschlag", "Es muss eine Datei im Format xls oder xlsx angegeben werden")
    sys.exit()
if excel_file1_path.endswith('.xls'):
    excel_file1_path = convert_to_xlsx(excel_file1_path)

 # second upload
excel_file2_path = get_file_path()
if excel_file2_path.endswith(".xls") or excel_file2_path.endswith(".xlsx"):
    info_window("Erfolg", "Die zweite Excel-Datei wurde angenommen!")
else:
    info_window("Fehlschlag", "Es muss eine Datei im Format xls oder xlsx angegeben werden")
    sys.exit()
if excel_file2_path.lower().endswith('.xls'):
    excel_file2_path = convert_to_xlsx(excel_file2_path)

df1 = pd.read_excel(excel_file1_path) # Read both Excel files into pandas dataframes
df2 = pd.read_excel(excel_file2_path)

common_numbers = pd.merge(df1, df2, how='inner', left_on=df1.columns[0], right_on=df2.columns[0]) # Find common numbers in the first column

color_overlapping_numbers(excel_file1_path, common_numbers[df1.columns[0]]) # Color overlapping numbers in the first selected Excel file
color_overlapping_numbers(excel_file2_path, common_numbers[df2.columns[0]]) # Color overlapping numbers in the second selected Excel file
info_window("Fertig", "Deine markierten Excel Dateien findest du unter\n{}\nund\n{}.".format(excel_file1_path,excel_file2_path))
