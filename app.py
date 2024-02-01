import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def get_file_path():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title="Excel-Datei auswählen")
    return file_path

def convert_to_xlsx(file_path):
    # Read the .xls file into a DataFrame
    df = pd.read_excel(file_path)

    # Save the DataFrame to a new .xlsx file
    df.to_excel(file_path + 'x', index=False)
    return file_path + 'x'

def color_overlapping_numbers(excel_file_path, common_numbers):
    # Load the Excel file using openpyxl
    wb = load_workbook(excel_file_path)

    # Access the active sheet
    ws = wb.active

    # Get the column letter for the first column
    col_letter = ws.cell(row=1, column=1).column_letter

    # Set the fill color for overlapping numbers
    for idx, value in enumerate(pd.read_excel(excel_file_path)[pd.read_excel(excel_file_path).columns[0]].isin(common_numbers)):
        if value:
            cell = ws[col_letter + str(idx + 2)]  # Adding 2 because Excel indexing starts from 1
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow fill color

    # Save the modified workbook
    wb.save(excel_file_path)

# Prompt user to upload the first Excel file
print("Wählen Sie die erste Excel-Datei aus:")
excel_file1_path = get_file_path()

# Convert to .xlsx if the file is in .xls format
if excel_file1_path.lower().endswith('.xls'):
    excel_file1_path = convert_to_xlsx(excel_file1_path)

# Prompt user to upload the second Excel file
print("Wählen Sie die zweite Excel-Datei aus:")
excel_file2_path = get_file_path()

# Convert to .xlsx if the file is in .xls format
if excel_file2_path.lower().endswith('.xls'):
    excel_file2_path = convert_to_xlsx(excel_file2_path)

# Read both Excel files into pandas dataframes
df1 = pd.read_excel(excel_file1_path)
df2 = pd.read_excel(excel_file2_path)

# Find common numbers in the first column
common_numbers = pd.merge(df1, df2, how='inner', left_on=df1.columns[0], right_on=df2.columns[0])

# Color overlapping numbers in the first selected Excel file
color_overlapping_numbers(excel_file1_path, common_numbers[df1.columns[0]])

# Color overlapping numbers in the second selected Excel file
color_overlapping_numbers(excel_file2_path, common_numbers[df2.columns[0]])